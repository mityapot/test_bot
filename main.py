import config as conf
from openpyxl import load_workbook
from pymongo import MongoClient
import telebot
from threading import Thread

bot = telebot.TeleBot(conf.TOKEN)


@bot.message_handler(commands=['start'])
def welcome_start(message):
    """
    Команда приветствия бота
    """
    bot.send_message(message.chat.id, 'Приветствую, для выгрузки фильмов введите команду /view')


@bot.message_handler(commands=['view'])
def view_films(message):
    """
    Выгрузка данных из базы данных в бот
    """
    with client:
        db = client[conf.DB]
        all_films = db.films.find()
        str_all = []
        for film in all_films:
            str_all.append("{} - {} - {} - {}".format(film['name'], film['director'], film['year'], film['views']))
    bot.send_message(message.chat.id, "Данные в формате: Название фильма - Режиссер - Год - Количество просмотров")
    if len(str_all) != 0:
        bot.send_message(message.chat.id, "\n".join(str_all))


def read_data(client):
    """
    Чтение excel и занесение в базу данных
    """
    wb = load_workbook('data.xlsx')
    sheet = wb.active
    row_count = sheet.max_row
    new_data = []
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=row_count, max_col=4):
        new_data.append({"name": row[0].value, "year": row[1].value, "director": row[2].value, "views": row[3].value})
    with client:
        db = client[conf.DB]
        films_collection = db['films']
        films = films_collection.insert_many(new_data)
    print(films.inserted_ids)


if __name__ == '__main__':
    client = MongoClient(conf.HOST, conf.PORT)
    th = Thread(target=bot.polling)
    th.start()
    cmd = input("Введите команду: ")
    while cmd != "0":
        if cmd == "1":
            read_data(client)
        cmd = input("Введите команду: ")
    bot.stop_polling()
    th.join()
